import os
import re
from PIL import Image
import customtkinter
from tkinter import filedialog, END
from openpyxl.styles import Font
from openpyxl import Workbook
import openpyxl
import tkinter.messagebox as messagebox

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
#создание базового интерфейса
        self.title("ExcelReader")
        self.geometry("700x450")
        self.iconbitmap("logo.ico")
        self.resizable(False, False)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)

        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.second_frame.grid_columnconfigure(0, weight=1)

        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.create_table_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.create_table_frame.grid_columnconfigure(0, weight=1)

        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join("logo.png")), size=(24, 24))

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text="ExcelReader",
                            compound="left", font=customtkinter.CTkFont(size=15, weight="bold"), image=self.logo_image)
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Главная",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Настройки",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["темная тема", "светлая тема"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")

        self.home_frame_large_image_label = customtkinter.CTkLabel(self.home_frame, text="")
        self.home_frame_large_image_label.grid(row=0, column=0, padx=20, pady=10)

        self.label_selection = customtkinter.CTkLabel(self.home_frame, text="Выберите операцию")
        self.label_selection.grid(row=0, column=1, padx=14, pady=10, columnspan=4)

        self.create_table_button = customtkinter.CTkButton(self.home_frame, text="Создание", corner_radius=5,
                                                            height=40, command=self.create_table_operation)
        self.create_table_button.grid(row=1, column=1, padx=13, pady=10)

        self.overwrite_table_button = customtkinter.CTkButton(self.home_frame, text="Перезаписать/Сохранить",
                                                               corner_radius=5, height=40, command=self.overwrite_table_operation)
        self.overwrite_table_button.grid(row=1, column=2, padx=13, pady=10)

        self.delete_table_button = customtkinter.CTkButton(self.home_frame, text="Удаление таблицы",
                                                            corner_radius=5, height=40, command=self.delete_table_frame)
        self.delete_table_button.grid(row=1, column=3, padx=13, pady=10)

        self.label_shrift = customtkinter.CTkLabel(self.second_frame,text="Разделитель для выделения шрифта")
        self.label_shrift.grid(row=0, column=0, padx=20, pady=10)

        self.entry_shrift = customtkinter.CTkEntry(self.second_frame)
        self.entry_shrift.grid(row=1, column=0, padx=20, pady=10)

        self.shrift_button = customtkinter.CTkButton(self.second_frame, text="Сохранить настройки", corner_radius=5,
                                                            height=40, command=self.save_settings)
        self.shrift_button.grid(row=2, column=0, padx=20, pady=10)

        self.load_settings()
        self.select_frame_by_name("home")

    def load_settings(self): #функция для загруки txt файла
        try:
            with open("settings.txt", "r") as file:
                separator = file.read()

                self.entry_shrift.delete(0, "end")

                self.entry_shrift.insert("end", separator)
        except FileNotFoundError:
            pass

    def save_settings(self): #функция для сохранения txt файла
        separator = self.entry_shrift.get()
        with open("settings.txt", "w") as file:
            file.write(separator)
        messagebox.showinfo("Успех",f'Настройки сохранены"')

    def create_table_operation(self): #интерфейс для операции в кнопке Создание
        self.table_frame_operatin = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.table_frame_operatin.grid_columnconfigure(0, weight=1)

        save_sheet = customtkinter.CTkButton(self.table_frame_operatin, text="Создать таблицу", corner_radius=5, height=40,
                                              command=self.create_table_frane)
        save_sheet.grid(row=0, column=0, padx=10, pady=10)

        save_table = customtkinter.CTkButton(self.table_frame_operatin, text="Создать лист", corner_radius=5, height=40,
                                              command=self.create_list_frane)
        save_table.grid(row=1, column=0, padx=10, pady=10)


        self.table_frame_operatin.grid(row=0, column=1, sticky="nsew")

    def overwrite_table_operation(self): #интерфейс для операций перезаписи
        self.overwrite_frame_operatin = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.overwrite_frame_operatin.grid_columnconfigure(0, weight=1)

        overwrite_info = customtkinter.CTkButton(self.overwrite_frame_operatin, text="Перезаписать информацию", corner_radius=5, height=40,
                                              command=self.overwrite_table_frane)
        overwrite_info.grid(row=0, column=0, padx=10, pady=10)

        overwrite_table = customtkinter.CTkButton(self.overwrite_frame_operatin, text="Перезаписать название таблицы", corner_radius=5, height=40,
                                              command=self.overqite_table_reloaded)
        overwrite_table.grid(row=1, column=0, padx=10, pady=10)


        self.overwrite_frame_operatin.grid(row=0, column=1, sticky="nsew")

    def overwrite_table_frane(self): #интерфейс для работы с перезаписью
        self.overwrite_info_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.overwrite_info_frame.grid_columnconfigure(0, weight=1)

        self.sheet_entry = customtkinter.CTkEntry(self.overwrite_info_frame, placeholder_text="Название листа...")
        self.sheet_entry.grid(row=1, column=0, padx=20, pady=10)

        self.browse_button = customtkinter.CTkButton(self.overwrite_info_frame, text="Указать путь к файлу", corner_radius=0, height=40,
                                                      command=self.browse_file)

        self.browse_button.grid(row=3, column=0, padx=20, pady=10)

        self.overwrite_button = customtkinter.CTkButton(self.overwrite_info_frame, text="Перезаписать информацию",
                                                        corner_radius=0, height=40, command=self.overwrite_file)
        self.overwrite_button.grid(row=5, column=0, padx=20, pady=10)

        self.text_widget = customtkinter.CTkTextbox(self.overwrite_info_frame)
        self.text_widget.grid(row=4, column=0, padx=20, pady=10, sticky="nsew")

        self.overwrite_info_frame.grid(row=0, column=1, sticky="nsew")

    def create_table_frane(self): #интерфейс для создания таблицы
        self.table_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.table_frame.grid_columnconfigure(0, weight=1)

        label = customtkinter.CTkLabel(self.table_frame, text="Новая таблица")
        label.grid(row=0, column=0, padx=20, pady=10)

        self.entry_name_table  = customtkinter.CTkEntry(self.table_frame, placeholder_text="Название таблицы...")
        self.entry_name_table.grid(row=1, column=0, padx=20, pady=10)

        save_button = customtkinter.CTkButton(self.table_frame, text="Создать", corner_radius=0, height=40,
                                              command=self.create_table)
        save_button.grid(row=2, column=0, padx=20, pady=10)

        self.table_frame.grid(row=0, column=1, sticky="nsew")

    def return_to_home_table(self):
        self.select_frame_by_name("home")

    def create_table(self): #функция для создания таблицы
        table_name = self.entry_name_table.get()

        if not table_name:
            messagebox.showwarning("Предупреждение", "Введите название таблицы")
            return

        workbook = Workbook()

        file_path = f"{table_name}.xlsx"
        workbook.save(file_path)

        messagebox.showinfo("Успешно", f"Таблица '{table_name}' успешно создана и сохранена.")
        self.return_to_home()

    def create_list_frane(self):#интерфейс для создания листа
        self.list_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.list_frame.grid_columnconfigure(0, weight=1)

        self.entry_name_list = customtkinter.CTkEntry(self.list_frame, placeholder_text="Название листа...")
        self.entry_name_list.grid(row=1, column=0, padx=20, pady=10)

        list_button = customtkinter.CTkButton(self.list_frame, text="Выбор таблицы", corner_radius=0, height=40,
                                              command=self.create_list)
        list_button.grid(row=2, column=0, padx=20, pady=10)

        self.list_frame.grid(row=0, column=1, sticky="nsew")

    def overqite_table_reloaded(self): #интерфейс для перезаписи названия таблицы
        self.over_write = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.over_write.grid_columnconfigure(0, weight=1)

        self.over_name_table = customtkinter.CTkEntry(self.over_write, placeholder_text="Новое название таблицы...")
        self.over_name_table.grid(row=1, column=0, padx=20, pady=10)

        peremen = customtkinter.CTkButton(self.over_write, text="Переименовать", corner_radius=0, height=40,
                                              command=self.rename_table)
        peremen.grid(row=3, column=0, padx=20, pady=10)


        self.over_write.grid(row=0, column=1, sticky="nsew")

    def rename_table(self):
        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            new_table_name = self.over_name_table.get()

            if file_path and new_table_name:
                directory, old_filename = os.path.split(file_path)
                new_filename = os.path.join(directory, f'{new_table_name}{os.path.splitext(old_filename)[1]}')

                os.rename(file_path, new_filename)

                messagebox.showinfo("Успех",
                                    f'Файл переименован в "{new_table_name}{os.path.splitext(old_filename)[1]}"')
                self.return_to_home()
            else:
                messagebox.showwarning("Предупреждение", 'Пожалуйста, выберите файл и введите новое имя.')
        except Exception as e:
            messagebox.showerror("Ошибка", f'Ошибка: {str(e)}')
    def create_list(self): #функция для создания листа
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel files", "*.xlsx;*.xls")])
        workbook = openpyxl.load_workbook(file_path)
        sheet = self.entry_name_list.get()
        new_sheet = workbook.create_sheet(sheet)
        workbook.save(file_path)
        messagebox.showinfo("Успех",f'Лист создан"')
        self.return_to_home()

    def return_to_home(self):
        self.home_frame.grid_forget()
        self.select_frame_by_name("home")

    def delete_table_frame(self): #интерфейс для удаления таблиц
        self.table_delete_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.table_delete_frame.grid_columnconfigure(0, weight=1)

        save_button = customtkinter.CTkButton(self.table_delete_frame, text="Удалить", corner_radius=0, height=40,
                                              command=self.delete_table)
        save_button.grid(row=2, column=0, padx=20, pady=10)

        self.table_delete_frame.grid(row=0, column=1, sticky="nsew")

        self.return_to_home()


    def delete_table(self): #функция для удаления таблиц
        excel_file_path = filedialog.askopenfilename(title="Выберите файл Excel",
                                                     filetypes=[("Excel files", "*.xlsx;*.xls")])

        if excel_file_path:
            try:
                os.remove(excel_file_path)
                print(f"Файл '{os.path.basename(excel_file_path)}' удален.")
                messagebox.showinfo("Успешно", f"Таблица '{excel_file_path}' успешно удалена.")
            except Exception as e:
                print(f"Произошла ошибка при удалении файла: {e}.")
            self.return_to_home()

        self.text_widget = customtkinter.CTkTextbox(self.home_frame)
        self.text_widget.grid(row=2, column=0, padx=20, pady=10, sticky="nsew")


    def select_frame_by_name(self, name): #функция для перекллючения между главной и настройками
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")

        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.second_frame.grid_forget()

    def home_button_event(self): #фунция для переклбчения на главную с других фреймов
        for children in self.winfo_children():
            if isinstance(children, customtkinter.CTkFrame):
                if children.winfo_ismapped():
                    children.grid_forget()
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def change_appearance_mode_event(self, new_appearance_mode): #функция для изменения внешнего вида программы
        if new_appearance_mode == "темная тема":
            customtkinter.set_appearance_mode("Dark")
        else:
            customtkinter.set_appearance_mode("Light")

    def browse_file(self): #функция для вывода данных из excel файла
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            content = "\n".join(
                [",".join([f"{cell.value if cell.value is not None else ''}" for cell in row if cell.value is not None])
                 for row in worksheet.iter_rows()])

            self.text_widget.delete(1.0, "end")
            self.text_widget.insert("end", content)

    def overwrite_file(self): #функция для перезаписи данных
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

        if file_path:
            workbook = openpyxl.load_workbook(file_path)

            sheet_name = self.sheet_entry.get()

            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.value = None
                        cell.font = None 

                start_row = 1
                start_column = 1

                new_content = self.text_widget.get("1.0", END).strip().split('\n')
                markup = re.compile(re.escape(self.entry_shrift.get()))
                for row_index, row in enumerate(new_content):
                    data = row.strip().split(',')
                    for column_index, value in enumerate(data):
                        value_without_markup = markup.sub('', value)
                        cell = worksheet.cell(row=start_row + row_index, column=start_column + column_index,
                                              value=value_without_markup)
                        if markup.search(value):
                            cell.font = Font(bold=True)
                workbook.save(file_path)
                messagebox.showinfo("Успешно", "Данные успешно добавлены в файл Excel.")
            else:
                messagebox.showerror("Ошибка", f"Лист {sheet_name} не существует.")

            self.return_to_home()

app = App()
app.load_settings()
app.mainloop()